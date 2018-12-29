from openpyxl import load_workbook#读取excel数据的第三方库
import logging
import datetime

class DoExcel:
    def __init__(self,excelpath):
        #打开excel方法
        self.wb=load_workbook(excelpath)
        self.sh_case_data=self.wb["case_datas"]
        self.sh_prepare_data = self.wb["prepare_datas"]

        # 获取当前时间+1小时的开始时间，当前时间+2小时的结束时间
    def get_init_datas(self):
        init_datas = {}
        for index in range(2, self.sh_prepare_data.max_row + 1):
            key = self.sh_prepare_data.cell(row=index, column=1).value
            init_datas[key] = str(self.sh_prepare_data.cell(row=index, column=2).value)
        init_datas["${begintime}"] = str(init_datas["${begintime}"])
        init_datas["${finishtime}"] = str(init_datas["${finishtime}"])
        #logging.info("初始化数据为：{0}".format(init_datas))
        return init_datas

        # 读取并初始化数据
    def get_caseDatas_all(self):
        all_case_datas = []
        #读取excel数据规则：从第二行的第一列开始（这里的下标是从1开始的）
        #for读取数据的规则for i in range（2,3）循环的是2
        for index in range(2,self.sh_case_data.max_row+1):
            case_data={}
            #读取出来的数据放在数组中，一一对应
            case_data["case_id"]=self.sh_case_data.cell(row=index,column=1).value#用例编号
            case_data["shuoming"]=self.sh_case_data.cell(row=index,column=3).value#用例说明
            case_data["api_name"]=self.sh_case_data.cell(row=index,column=4).value#接口名称
            case_data["method"]=self.sh_case_data.cell(row=index,column=5).value#请求接口的方式get还是post
            case_data["url"]=self.sh_case_data.cell(row=index,column=6).value#请求的url
            case_data["request_data"] = self.sh_case_data.cell(row=index, column=7).value#请求的参数
            case_data["compare_type"] = self.sh_case_data.cell(row=index, column=9).value  # 是否需要正则匹配的参数
            temp_case_data = str(self.sh_case_data.cell(row=index, column=10).value) # 开始时间和结束时间需要用参数替换的
            # 获取初始值
            init_datas = self.get_init_datas()
            #print(init_datas)
            if temp_case_data is not None and len(init_datas) > 0:
                for key, value in init_datas.items():
                    # 如果找到了，则替换。
                    if case_data["request_data"].find(key) != -1:
                        case_data["request_data"] = case_data["request_data"].replace(key, value)
            #logging.info("初始化之后的请求数据为：\n{0}".format(temp_case_data))
            case_data["request_data"] = case_data["request_data"]
            case_data["expected_data"]=self.sh_case_data.cell(row=index,column=8).value#获取期望数据
            # 判断本用例是否需要对响应结果进行解析，并获取其中的值。是否需要进行正则匹配
            '''
            if self.sh_case_data.cell(row=index,column=9).value is not None:
                case_data["related_exp"]=self.sh_case_data.cell(row=index,column=10).value
            #print(case_data)
            '''
            all_case_datas.append(case_data)#新增数据值
        return all_case_datas

    # 将开始时间和结束时间写入excel中
    def write_BtFt(self):
        self.sh_prepare_data.cell(row=2, column=2).value = (datetime.datetime.now() + datetime.timedelta(hours=1)).strftime("%Y-%m-%d %H:%M:%S")  # 开始时间
        self.sh_prepare_data.cell(row=3, column=2).value = (datetime.datetime.now() + datetime.timedelta(hours=2)).strftime("%Y-%m-%d %H:%M:%S")  # 结束时间
    #保存数据
    def save_excelFile(self,excelpath):
        try:
            self.wb.save(excelpath)
        except Exception as e:
            logging.exception(e)
